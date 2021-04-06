import os
import sys
import openpyxl
import datetime
from openpyxl.styles import PatternFill, colors
import get_node_list

PATH = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PATH)
from settings.config import node_info_file

t = datetime.datetime.now().strftime('%Y-%m-%d')


def write_excel(node_info_dict, down_node_list):
    file_name = node_info_file
    workbook = openpyxl.load_workbook(file_name)
    new_sheet = workbook.create_sheet(t)
    sheet = workbook['模板']
    max_rows = sheet.max_row
    max_cols = sheet.max_column

    ## copy 模板
    for row in range(1, max_rows + 1):
        for col in range(1, max_cols + 1):
            data = sheet.cell(row=row, column=col).value
            new_sheet.cell(row=row, column=col, value=data)

    ## add node inof
    index = 1
    for row in range(2, max_rows + 1):
        if index == 1:
            new_sheet.cell(1, (max_cols + 1), 'core_num')
            new_sheet.cell(1, (max_cols + 2), 'max_used_cpu')
            new_sheet.cell(1, (max_cols + 3), 'min_used_cpu')
            new_sheet.cell(1, (max_cols + 4), 'avg_used_cpu')
            new_sheet.cell(1, (max_cols + 5), 'total_mem')
            new_sheet.cell(1, (max_cols + 6), 'max_used_mem')
            new_sheet.cell(1, (max_cols + 7), 'min_used_mem')
            new_sheet.cell(1, (max_cols + 8), 'avg_used_mem')
        index += 1
        ip = new_sheet.cell(row, 5).value
        if ip in node_info_dict.keys():
            new_sheet.cell(index, (max_cols + 1), node_info_dict[ip]['core_num'])
            new_sheet.cell(index, (max_cols + 2), node_info_dict[ip]['max_used_cpu'])
            new_sheet.cell(index, (max_cols + 3), node_info_dict[ip]['min_used_cpu'])
            new_sheet.cell(index, (max_cols + 4), node_info_dict[ip]['avg_used_cpu'])
            new_sheet.cell(index, (max_cols + 5), node_info_dict[ip]['total_mem'])
            new_sheet.cell(index, (max_cols + 6), node_info_dict[ip]['max_used_mem'])
            new_sheet.cell(index, (max_cols + 7), node_info_dict[ip]['min_used_mem'])
            new_sheet.cell(index, (max_cols + 8), node_info_dict[ip]['avg_used_mem'])
        else:
            pass

    ## down_node_list
    all_node_ip = down_node_list
    fill = PatternFill("solid", fgColor="1874CD")
    for row in range(1, max_rows):
        data = new_sheet.cell(row=row, column=5).value
        if data in all_node_ip:
            new_sheet.cell(row=row, column=5).fill = fill
            all_node_ip.remove(data)

    workbook.save(node_info_file)
